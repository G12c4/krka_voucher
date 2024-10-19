from datetime import date
from openpyxl import load_workbook
from pathlib import Path
from typing import List, Union
from rich import print
import os

def update_cells(ws, nguest_num: int, start_row: int, data: List[Union[str, int]], col: int) -> None:
    for num, value in enumerate(data, 1):
        position = ws.cell(nguest_num, col)
        position.value = value
        nguest_num += 1
    for _ in range(nguest_num, start_row, 1):
        position = ws.cell(nguest_num, col)
        position.value = ""
        nguest_num += 1

def write_data_to_excel(ticket_number: Union[str, List[str]], datum_in: str, ulaz_in: str, izlaz_in: str, ime_in: List[str], gr_in: List[str]) -> None:
    root_path =  Path(__file__).parent
    excel_path = root_path / "Vaucer_za_npkrka.xlsx"
    output_excel_dir = root_path.parents[1] / "Izdani_voucheri"

    now = date.today()
    year = now.strftime("%Y")
    wb = load_workbook(excel_path)
    ws = wb["Sheet1"]

    orig_voucher_num = ws.cell(12, 7)
    new_voucher_num = int(orig_voucher_num.value.split("/")[0]) + 1
    orig_voucher_num.value = f"{new_voucher_num}/{year}"

    ticket_number = str(ticket_number[0]) if isinstance(ticket_number, list) else str(ticket_number)
    ws.cell(19, 7).value = ticket_number

    ws.cell(20, 4).value = datum_in
    ws.cell(18, 1).value = f"ULAZ {ulaz_in if ulaz_in else 'LOZOVAC'},"
    ws.cell(18, 3).value = f"IZLAZ {izlaz_in if izlaz_in else 'LOZOVAC'}"

    update_cells(ws, 24, 34, [f"{num + 1}." for num in range(len(ime_in))], 1)
    update_cells(ws, 24, 34, ime_in, 3)
    update_cells(ws, 24, 34, gr_in, 6)

    ws.cell(36, 2).value = now.strftime("%d.%m.%Y")

    output_excel_path = str(output_excel_dir / f"Vaucer_za_npkrka_0{new_voucher_num}-{year}.xlsx")
    print(output_excel_path)

    wb.save(excel_path)
    wb.save(output_excel_path)
    
    os.startfile(output_excel_path, 'print')
    return True

def main():
    ticket_number = ["000000000001-Test"]
    datum_in = "01/04/2024"
    ulaz_in = "LOZOVAC"
    izlaz_in = "LOZOVAC"
    ime_in = ["Tina Sladić", "Josip Grcić"]
    gr_in = ["22/06/1986", "20/12/1990"]
    pdf_path = write_data_to_excel(ticket_number, datum_in, ulaz_in, izlaz_in, ime_in, gr_in)
    print(pdf_path)

if __name__ == '__main__':
    main()